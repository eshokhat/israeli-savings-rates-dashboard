"""
export_excel.py
---------------
Runs all analytical queries against savings.duckdb and writes each result
to a separate sheet in savings_analysis.xlsx.

Intended audience: non-technical stakeholders ("management", other teams).
The Tableau dashboard is built separately from the CSV.

Usage:
    python pipeline/export_excel.py
    python pipeline/export_excel.py --db data/savings.duckdb --out excel/savings_analysis.xlsx
"""

import argparse
import logging
from pathlib import Path

import duckdb
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DEFAULT_DB = Path("data/savings.duckdb")
DEFAULT_OUT = Path("excel/savings_analysis.xlsx")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ── style constants ───────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", start_color="1F4E79")  # dark blue
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
ROW_FONT = Font(name="Arial", size=10)
ALT_FILL = PatternFill("solid", start_color="EBF3FB")  # light blue stripe
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
HEADER_HEIGHT = 30
ROW_HEIGHT = 18

# ── queries ───────────────────────────────────────────────────────────────────
# Each entry: (sheet_name, sql, description_for_A1_note)
QUERIES: list[tuple[str, str, str]] = [
    (
        "Best Rates Today",
        """
        SELECT BANK_EN AS Bank, SAVINGSPLAN_EN AS Plan,
               SAVINGSPROGRAMBYAGE_EN AS "Age Group", RateType AS "Rate Type",
               ROUND(LatestRate, 4) AS "Latest Rate (%)", AsOfDate AS "As of Date"
        FROM vw_best_rates
        ORDER BY "Latest Rate (%)" DESC
        """,
        "Best available rate per bank, plan and age group — most recent data point.",
    ),
    (
        "Avg Rate by Bank & Plan",
        """
        SELECT BANK_EN          AS Bank,
               SAVINGSPLAN_EN   AS Plan,
               RateType         AS "Rate Type",
               ROUND(AVG(RateValue), 4) AS "Avg Rate (%)",
               COUNT(*)                 AS Observations
        FROM savings_rates
        WHERE RateValue IS NOT NULL
        GROUP BY BANK_EN, SAVINGSPLAN_EN, RateType
        ORDER BY BANK_EN, SAVINGSPLAN_EN, RateType
        """,
        "Average interest rate across full history, grouped by bank, plan type and rate type.",
    ),
    (
        "Under 18 vs Above 18",
        """
        WITH age_pivot AS (
            SELECT BANK_EN, SAVINGSPLAN_EN, RateType,
                   ROUND(AVG(RateValue) FILTER (
                       WHERE SAVINGSPROGRAMBYAGE_EN = 'Under 18'), 4) AS avg_u18,
                   ROUND(AVG(RateValue) FILTER (
                       WHERE SAVINGSPROGRAMBYAGE_EN = 'Above 18'), 4) AS avg_a18
            FROM savings_rates
            WHERE RateValue IS NOT NULL
            GROUP BY BANK_EN, SAVINGSPLAN_EN, RateType
        )
        SELECT BANK_EN          AS Bank,
               SAVINGSPLAN_EN   AS Plan,
               RateType         AS "Rate Type",
               avg_u18          AS "Avg Under 18 (%)",
               avg_a18          AS "Avg Above 18 (%)",
               ROUND(avg_u18 - avg_a18, 4) AS "Gap (Under - Above)"
        FROM age_pivot
        WHERE avg_u18 IS NOT NULL
           OR avg_a18 IS NOT NULL
        ORDER BY ABS(COALESCE(avg_u18, 0) - COALESCE(avg_a18, 0)) DESC
        """,
        "Difference in average rates between children's (Under 18) and adult (Above 18) savings programs.",
    ),
    (
        "Volatility by Bank",
        """
        WITH base AS (
            SELECT
                BANK_EN,
                RateType,
                ROUND(MAX(RateValue) - MIN(RateValue), 4)  AS Delta,
                ROUND(STDDEV(RateValue),               4)  AS StdDev,
                ROUND(MIN(RateValue),                  4)  AS MinRate,
                ROUND(MAX(RateValue),                  4)  AS MaxRate,
                ROUND(AVG(RateValue),                  4)  AS AvgRate,
                COUNT(*)                                   AS Observations
            FROM savings_rates
            WHERE RateValue IS NOT NULL
            GROUP BY BANK_EN, RateType
        ),
        ranked AS (
            SELECT *,
                   CASE NTILE(3) OVER (PARTITION BY RateType ORDER BY Delta ASC)
                       WHEN 1 THEN 'Stable'
                       WHEN 2 THEN 'Moderate'
                       ELSE 'Volatile'
                   END AS Stability
            FROM base
        )
        SELECT BANK_EN   AS Bank,
               RateType  AS "Rate Type",
               Delta      AS "Δ (Max–Min)",
               StdDev     AS "Std Dev",
               MinRate    AS "Min Rate (%)",
               MaxRate    AS "Max Rate (%)",
               AvgRate    AS "Avg Rate (%)",
               Observations,
               Stability
        FROM ranked
        ORDER BY "Rate Type", "Δ (Max–Min)" ASC
        """,
        "Rate volatility (max–min range) per bank. Stable = smallest Δ, Volatile = largest.",
    ),
    (
        "Monthly Trends",
        """
        WITH monthly AS (
            SELECT
                YEAR_MONTH,
                MIN(INTERESTSDATE) AS PeriodStart,
                RateType,
                ROUND(AVG(RateValue), 4) AS AvgRate
            FROM savings_rates
            WHERE RateValue IS NOT NULL
            GROUP BY YEAR_MONTH, RateType
        ),
        with_lag AS (
            SELECT
                YEAR_MONTH,
                RateType,
                AvgRate,
                ROUND(AvgRate - LAG(AvgRate) OVER (
                    PARTITION BY RateType ORDER BY PeriodStart
                ), 4) AS MoM_Delta
            FROM monthly
        )
        SELECT
            YEAR_MONTH      AS "Year-Month",
            RateType        AS "Rate Type",
            AvgRate         AS "Avg Rate (%)",
            MoM_Delta       AS "MoM Δ",
            CASE
                WHEN MoM_Delta > 0 THEN 'Rising'
                WHEN MoM_Delta < 0 THEN 'Falling'
                WHEN MoM_Delta = 0 THEN 'Stable'
                ELSE 'n/a'
            END             AS Trend
        FROM with_lag
        ORDER BY "Rate Type", "Year-Month"
        """,
        "Month-over-month average rate change per rate type. MoM Δ = current minus previous month.",
    ),
]


# ── formatting helpers ────────────────────────────────────────────────────────
def _write_sheet(ws, rows: list[list], headers: list[str], description: str) -> None:
    """Write headers + data rows with consistent formatting."""
    # Row 1: description note
    ws.append([description])
    desc_cell = ws.cell(row=1, column=1)
    desc_cell.font = Font(name="Arial", size=9, italic=True, color="595959")
    desc_cell.alignment = LEFT
    ws.row_dimensions[1].height = 24

    # Row 2: blank spacer
    ws.append([])

    # Row 3: headers
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
    ws.row_dimensions[3].height = HEADER_HEIGHT

    # Rows 4+: data
    for row_idx, row in enumerate(rows, start=4):
        ws.append(row)
        ws.row_dimensions[row_idx].height = ROW_HEIGHT
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = ROW_FONT
            cell.alignment = LEFT
            if fill:
                cell.fill = fill

    # Freeze pane below header row
    ws.freeze_panes = ws.cell(row=4, column=1)

    # Auto-fit column widths
    for col_idx, header in enumerate(headers, start=1):
        col_values = [str(header)] + [
            str(ws.cell(row=r, column=col_idx).value or "")
            for r in range(4, ws.max_row + 1)
        ]
        max_len = min(max(len(v) for v in col_values), 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3


def _add_cover_sheet(wb: Workbook, sheet_names: list[str]) -> None:
    """First sheet: index of all report sheets."""
    ws = wb.active
    ws.title = "Contents"

    ws.append(["Israeli Savings Rates — Analytical Report"])
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(name="Arial", size=14, bold=True, color="1F4E79")
    ws.row_dimensions[1].height = 30

    ws.append(["Generated automatically by export_excel.py from savings.duckdb"])
    ws.cell(row=2, column=1).font = Font(
        name="Arial", size=9, italic=True, color="595959"
    )
    ws.append([])

    ws.append(["Sheet", "Description"])
    for col_idx in range(1, 3):
        cell = ws.cell(row=4, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
    ws.row_dimensions[4].height = HEADER_HEIGHT

    descriptions = {q[0]: q[2] for q in QUERIES}
    for row_idx, name in enumerate(sheet_names, start=5):
        ws.cell(row=row_idx, column=1).value = name
        ws.cell(row=row_idx, column=2).value = descriptions.get(name, "")
        for col_idx in range(1, 3):
            ws.cell(row=row_idx, column=col_idx).font = ROW_FONT
            ws.cell(row=row_idx, column=col_idx).alignment = LEFT
        ws.row_dimensions[row_idx].height = ROW_HEIGHT

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 70


# ── main ─────────────────────────────────────────────────────────────────────
def export(db_path: Path, out_path: Path) -> None:
    con = duckdb.connect(str(db_path))
    wb = Workbook()

    sheet_names = [q[0] for q in QUERIES]
    _add_cover_sheet(wb, sheet_names)

    for sheet_name, sql, description in QUERIES:
        log.info("Running query → sheet '%s'", sheet_name)
        df = con.execute(sql.strip()).df()
        for col in df.select_dtypes(include=["datetime64[ns]", "datetimetz"]).columns:
            df[col] = df[col].dt.date
        headers = list(df.columns)
        rows = df.values.tolist()

        ws = wb.create_sheet(title=sheet_name)
        _write_sheet(ws, rows, headers, description)
        log.info("  %d rows written", len(rows))

    con.close()

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    log.info("Saved → %s", out_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Export DuckDB query results to Excel")
    parser.add_argument("--db", type=Path, default=DEFAULT_DB)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    args = parser.parse_args()

    if not args.db.exists():
        log.error("Database not found: %s — run load.py first", args.db)
        raise SystemExit(1)

    export(args.db, args.out)


if __name__ == "__main__":
    main()
